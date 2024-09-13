import pyperclip 
import openpyxl
import pyautogui

workbook = openpyxl.load_workbook('F:/Repo_GitHub/Codes_in_Py/auto preenchimento/produtos_ficticios.xlsx')
print(workbook.sheetnames)

sheet_produto = workbook['Produtos']

for linha in sheet_produto.iter_rows(min_row=2):
    nome_produto = linha[0].value
    pyperclip.copy(nome_produto)
    pyautogui.click(x=375, y=335,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    descricao_produto = linha[1].value
    pyperclip.copy(descricao_produto)
    pyautogui.click(x=366, y=434,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    categoria_produto = linha[2].value
    pyperclip.copy(categoria_produto)
    pyautogui.click(x=365, y=560,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    codigo_produto = linha[3].value
    pyperclip.copy(codigo_produto)
    pyautogui.click(x=383, y=651,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    peso = linha[4].value
    pyperclip.copy(peso)
    pyautogui.click(x=382, y=725,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    dimesoes = linha[5].value
    pyperclip.copy(dimesoes)
    pyautogui.click(x=416, y=820,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(x=351, y=869,duration=1)
    
    preco = linha[6].value
    pyperclip.copy(preco)
    pyautogui.click(x=456, y=367,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    quantidade = linha[7].value
    pyperclip.copy(quantidade)
    pyautogui.click(x=387, y=448,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    data_validade = linha[8].value
    pyperclip.copy(data_validade)
    pyautogui.click(x=409, y=528,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    cor = linha[9].value
    pyperclip.copy(cor)
    pyautogui.click(x=403, y=625,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    tamanho = linha[10].value
    pyautogui.click(x=371, y=710,duration=1)
    if tamanho == 'Pequeno':
        pyautogui.click(x=357, y=739,duration=1)
    elif tamanho == 'Médio':
        pyautogui.click(x=348, y=764,duration=1)
    else:
        pyautogui.click(x=377, y=781,duration=1) 

    
    material = linha[11].value
    pyperclip.copy(material)
    pyautogui.click(x=372, y=799,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(x=343, y=852,duration=1)
    
    fabricante = linha[12].value
    pyperclip.copy(fabricante)
    pyautogui.click(x=357, y=381,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    pais_origem = linha[13].value
    pyperclip.copy(pais_origem)
    pyautogui.click(x=365, y=467,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    observacoes = linha[14].value
    pyperclip.copy(observacoes)
    pyautogui.click(x=349, y=563,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    codigo_barras = linha[15].value
    pyperclip.copy(codigo_barras)
    pyautogui.click(x=393, y=689,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    localizacao_aramazem = linha[16].value
    pyperclip.copy(localizacao_aramazem)
    pyautogui.click(x=357, y=770,duration=1)
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(x=343, y=834,duration=1)
    pyautogui.click(x=1130, y=164,duration=1)
    pyautogui.click(x=934, y=616,duration=1)